from openpyxl import load_workbook


wb = load_workbook('sagatave_eksamenam.xlsx')
ws = wb['Lapa_0']


header_row = [cell.value for cell in ws[3]]
priority_idx = None
date_idx = None


for idx, col_name in enumerate(header_row):
    if col_name == "Prioritāte":
        priority_idx = idx
    if col_name == "Piegādes datums":
        date_idx = idx


if priority_idx is None or date_idx is None:
    print("Required columns 'Prioritāte' or 'Piegādes datums' not found.")
else:
    matching_count = 0

    for row in ws.iter_rows(min_row=4, values_only=True):
        priority = row[priority_idx]
        delivery_date = row[date_idx]

        if priority == "High" and hasattr(delivery_date, 'year') and delivery_date.year == 2015:
            matching_count += 1

    print("Number of entries with High priority and delivery year 2015:", matching_count)
