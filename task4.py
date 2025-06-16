from openpyxl import load_workbook
import math  # for rounding down

wb = load_workbook('sagatave_eksamenam.xlsx')
ws = wb['Lapa_0']

# Header row is at row 3
header_row = [cell.value for cell in ws[3]]
product_idx = None
price_idx = None


for idx, col_name in enumerate(header_row):
    if col_name == "Produkts":
        product_idx = idx
    if col_name == "Cena":
        price_idx = idx


if product_idx is None or price_idx is None:
    print("Required columns 'Produkts' or 'Cena' not found.")
else:
    total_price = 0
    count = 0

    for row in ws.iter_rows(min_row=4, values_only=True):
        product = row[product_idx]
        price = row[price_idx]

        if isinstance(product, str) and "LaserJet" in product:
            try:
                total_price += float(price)
                count += 1
            except (TypeError, ValueError):
                continue

    if count > 0:
        average = total_price / count
        print("Average price (rounded down):", math.floor(average))
    else:
        print("No matching LaserJet products found.")
