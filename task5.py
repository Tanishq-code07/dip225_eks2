from openpyxl import load_workbook
import math

# Load workbook with cached formula values
wb = load_workbook('sagatave_eksamenam.xlsx', data_only=True)
ws = wb['Lapa_0']

# Header row
header_row = [cell.value for cell in ws[3]]
client_idx = None
count_idx = None
total_idx = None

# Find needed column indexes
for idx, col_name in enumerate(header_row):
    if col_name == "Klients":
        client_idx = idx
    if col_name == "Skaits":
        count_idx = idx
    if col_name == "Kopā":
        total_idx = idx

if client_idx is None or count_idx is None or total_idx is None:
    print("Required columns not found.")
else:
    total_sum = 0

    for row in ws.iter_rows(min_row=4, values_only=True):
        client = row[client_idx]
        count = row[count_idx]
        total = row[total_idx]

        if client == "Korporatīvais":
            try:
                count_val = float(count)
                total_val = float(total)
                if 40 <= count_val <= 50:
                    total_sum += total_val
            except (TypeError, ValueError):
                continue

    print("Total sum (rounded down):", math.floor(total_sum))
