from openpyxl import load_workbook


wb = load_workbook('sagatave_eksamenam.xlsx')
ws = wb['Lapa_0']  # Sheet name is case-sensitive


header_row = [cell.value for cell in ws[3]]
address_col_idx = None
count_col_idx = None

for idx, col_name in enumerate(header_row):
    if col_name == "Adrese":
        address_col_idx = idx
    if col_name == "Skaits":
        count_col_idx = idx


if address_col_idx is None or count_col_idx is None:
    print("Required columns 'Adrese' or 'Skaits' not found.")
else:
    matching_records = 0
    matching_rows = []

    for row in ws.iter_rows(min_row=4, values_only=True):  
        address = row[address_col_idx]
        count = row[count_col_idx]

        if isinstance(address, str) and address.startswith("Ain"):
            try:
                if float(count) < 40:
                    matching_records += 1
                    matching_rows.append((address, count))
            except (TypeError, ValueError):
                continue

    
    print("Number of matching records:", matching_records)
    print("Matching rows:")
    for r in matching_rows:
        print(r)
