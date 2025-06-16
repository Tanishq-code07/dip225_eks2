from openpyxl import load_workbook

# Load workbook and the correct sheet
wb = load_workbook('sagatave_eksamenam.xlsx')
ws = wb['Lapa_0']

# Header is in row 3
header_row = [cell.value for cell in ws[3]]
address_idx = None
city_idx = None

# Identify column indices
for idx, col_name in enumerate(header_row):
    if col_name == "Adrese":
        address_idx = idx
    if col_name == "Pilsēta":
        city_idx = idx

# Check if both columns exist
if address_idx is None or city_idx is None:
    print("Required columns 'Adrese' or 'Pilsēta' not found.")
else:
    match_count = 0

    for row in ws.iter_rows(min_row=4, values_only=True):  # Data starts after header
        address = row[address_idx]
        city = row[city_idx]

        if address == "Adulienas iela" and city in ["Valmiera", "Saulkrasti"]:
            match_count += 1

    print("Number of entries with 'Adulienas iela' in Valmiera or Saulkrasti:", match_count)
